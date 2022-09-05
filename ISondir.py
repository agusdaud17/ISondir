"""
Package untuk mengolah data sondir berdasarkan Robertson 2009
Baik single point, maupun multi point test.
"""
import os
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.pyplot import figure
from matplotlib.ticker import FormatStrFormatter, ScalarFormatter
from scipy import signal, interpolate
from math import*
import numpy as np
from openpyxl import load_workbook

# from scipy import interpolate
from PIL import Image
import seaborn as sns
import matplotlib.patches as mpatches

import pondasi

im = np.array(Image.open('SBTlow.jpg'))
x_logref = np.linspace(-1,1,1649)
x_grid = np.arange(0,1649,1)


class Single:
    """docstring fo Sondir."""

    def __init__(self,id_="S-1",dz=0.2,a=0.8,qc_unit='MPa',fs_unit='kPa'):
        self.id_ = id_
        self.dz = dz
        self.a = a
        if qc_unit == 'MPa':
            self.qcunit = 1000
            self.to_kPa = 1000
        else:
            self.to_kPa = 100
            self.qcunit = 100

        if fs_unit == 'kPa':
            self.funit = 1
        else:
            self.funit = 100

    def from_excel(self, name=None, sheet="DATA (1)", colz='A',colqc='F',colfs='G',data_row=15,cell_date='D5', cell_loc='D6', cell_GWL='J7', cell_coord='D7'):
        workbook = load_workbook(filename=name,data_only=True)
        data_sheet = workbook[sheet]

        self.project_name = "-"
        self.date_test = data_sheet[cell_date].value
        self.location = data_sheet[cell_loc].value
        self.coordinate = data_sheet[cell_coord].value
        self.recordedby = "-"
        self.calculatedby = "-"
        self.checkedby = "-"
        GWL = data_sheet[cell_GWL].value

        if type(GWL) is str or None:
            self.GWL = 99
        else:
            self.GWL = GWL

        list_z = [0]
        list_qc = [0]
        list_fs = [0]

        i = data_row
        while True:
            qc = data_sheet[f'{colqc}{i}'].value
            if not type(qc) is float:
                data = list(zip(list_z,list_qc,list_fs))
                self.df = pd.DataFrame(data,columns = ['z', 'qc','fs'])
                self.df['u2'] = 0
                break
            elif qc > 0:
                list_z.append(data_sheet[f'{colz}{i}'].value)
                list_qc.append(data_sheet[f'{colqc}{i}'].value)
                list_fs.append(data_sheet[f'{colfs}{i}'].value)
                i+=1
            else:
                pass

    def zoneByColor(self,x,y):
        x_logref = np.linspace(-1,1,1649)
        x_grid = np.arange(0,1649,1)
        y_logref = np.linspace(0,3,1823)
        y_grid = np.arange(0,1823,1)
        if y > 1000:
            y = 1000
        if x > 10:
            x = 10
        fx = interpolate.interp1d(x_logref, x_grid)
        fy = interpolate.interp1d(y_logref, y_grid)
        xg = int(fx(log10(x)))
        yg = int(fy(log10(y)))

        while True:
            color = im[-yg,xg]
            if color[1] == 222:
                zone = 1
                break
            elif color[1] == 105:
                zone = 2
                break
            elif color[1] == 123 or color[1] == 122:
                zone = 3
                break
            elif color[1] == 83:
                zone = 4
                break
            elif color[1] == 148:
                zone = 5
                break
            elif color[1] == 248 or color[1] == 247 or color[1] == 249:
                zone = 6
                break
            elif color[1] == 160:
                zone = 7
                break
            elif color[1] == 102:
                zone = 8
                break
            elif color[1] == 33:
                zone = 9
                break
            else:
                yg -= 1
        return zone

    def read_excel(self,file_name):
        self.df = pd.read_excel(file_name)
        self.end_depth = min(self.df['z'])

    def data_frame(self,df):
        self.df = df
        self.end_depth = min(self.df['z'])

    def set_project(self,name='-'):
        self.project_name = name

    def set_date(self,date='-'):
        self.date_test = date

    def set_location(self,location='-'):
        self.location = location

    def set_coordinate(self,coord='-'):
        self.coordinate = coord

    def set_recordedby(self,name='-'):
        self.recordedby = name

    def set_calculatedby(self,name='-'):
        self.calculatedby = name

    def set_checkedby(self,name='-'):
        self.checkedby = name

    def average_qc(self):
        qc_filtered = []
        last_i = len(self.df['qc'])-1
        for i, qc in enumerate(self.df['qc']):
            if i == 0:
                qc_avg = 0
            elif i == last_i:
                qc_prev = self.df.loc[i-1,'qc']
                qc_current = self.df.loc[i,'qc']
                qc_avg = (qc_prev + qc_current) / 2
            else:
                qc_prev = self.df.loc[i-1,'qc']
                qc_current = self.df.loc[i,'qc']
                qc_forwad = self.df.loc[i+1,'qc']
                qc_avg = (qc_prev + qc_current + qc_forwad) / 3
            qc_filtered.append(qc_avg)
        self.df['qc_filtered'] = qc_filtered

    def average_fs(self):
        fs_filtered = []
        last_i = len(self.df['fs'])-1
        for i, fs in enumerate(self.df['fs']):
            if i == 0:
                fs_avg = 0
            elif i == last_i:
                fs_prev = self.df.loc[i-1,'fs']
                fs_current = self.df.loc[i,'fs']
                fs_avg = (fs_prev + fs_current) / 2
            else:
                fs_prev = self.df.loc[i-1,'fs']
                fs_current = self.df.loc[i,'fs']
                fs_forwad = self.df.loc[i+1,'fs']
                fs_avg = (fs_prev + fs_current + fs_forwad) / 3
            fs_filtered.append(fs_avg)
        self.df['fs_filtered'] = fs_filtered

    def qt_corrected(self,qc,u2):
        qt = round(qc + (u2*(1-self.a)),2)
        return qt

    def totalOverburdenStress(self,g):
        svo = g*self.dz
        return svo

    def normalizedFrictionRatio(self,fs,qt,svo):
        Rf = round((fs/(qt-svo))*100,2)
        return Rf

    def unitWeight(self,qt,Rf,gw=9.81,pa=100):
        g = round(((0.27*log10(Rf)) + (0.36*log10(qt/pa)) + 1.236)*gw,2)
        return max(g,15.5)

    def porePressure(self,gw):
        u = round(gw*self.dz,2)
        return u

    def effectiveOverburdenStress(svo,u):
        svo_ef = roun(svo - u,2)
        return svo_ef

    def normalizedConeResistance1990(self,qt,svo,svo_ef):
        Qt = round((qt-svo)/svo_ef,2)
        return Qt

    def normalizedConeResistance2009(self,qt,svo,svo_ef,n,pa=100):
        Qtn = round(((qt-svo)/pa)*(pa/svo_ef)**n,2)
        return Qtn

    def SBTindex(self,Qtn,Rf):
        Ic = round(sqrt(((3.47 - log10(Qtn))**2) + ((1.22+log10(Rf))**2)),2)
        return Ic

    def stressExponent(self,Ic,svo_ef,pa=100):
        n = round((0.381*Ic) + (0.05*svo_ef/pa) - 0.15,4)
        return min(n,1.0)

    def stepOne(self,svo_prev,qc,fs,u2):
        qt = self.qt_corrected(qc,u2)*self.qcunit # kPa
        fs = fs*self.funit # kPa
        Rf = (fs/qt)*100
        g = self.unitWeight(qt,Rf,gw=9.8,pa=100)
        svo = svo_prev + self.totalOverburdenStress(g)
        # Rf = self.normalizedFrictionRatio(fs,qt,svo)
        return round(g,2), round(svo,2), round(qt,2)

    def pc_mayne2012(self,m,qt,svo):
        pa = 100
        pc = 0.33 * ((qt-svo)**m) * (pa/100)**(1-m)
        return pc

    def solve_unitWeight(self):
        list_g = [0]
        list_svo = [0]
        list_qt = [0]

        for i in range(1,len(self.df['z'])):
            svo_prev = list_svo[i-1]
            qc = self.df.loc[i,'qc_filtered']
            fs = self.df.loc[i,'fs_filtered']
            u2 = self.df.loc[i,'u2']
            g, svo, qt= self.stepOne(svo_prev,qc,fs,u2)
            list_g.append(g)
            list_svo.append(svo)
            list_qt.append(qt)
        list_g[0] = list_g[1]
        self.df['g'] = list_g
        self.df['qt'] = list_qt

    def solve_RfSvo(self):
        list_svo = [0]
        list_Rf = [0]

        for i, g in enumerate(self.df['g']):
            if i > 0:
                svo_prev = list_svo[i-1]
                fs = self.df.loc[i,'fs_filtered']*self.funit
                qt = self.df.loc[i,'qt']
                svo = svo_prev + self.totalOverburdenStress(g=g)
                Rf = self.normalizedFrictionRatio(fs,qt,svo)
                list_svo.append(svo)
                list_Rf.append(Rf)
            else:
                pass
        self.df['svo'] = list_svo
        self.df['Rf'] = list_Rf

    def effectiveParams(self,z,gw=9.81):
        if z == None or self.GWL == None:
            u = 0
        elif z < abs(self.GWL):
            u = 0
        else:
            z = z-abs(self.GWL)
            u = z*gw
        return u

    def setGWL(self,z=None):
        self.GWL = z

    def solve_porePressure(self):
        list_u = []
        for z in self.df['z']:
            u = self.effectiveParams(abs(z))
            list_u.append(round(u,2))
        self.df['u']=list_u
        self.df['svo_ef'] = self.df['svo']-self.df['u']

    def SBTn(self,qt,svo,svo_ef,Rf,pa=100):
        n_init = 0.1

        while True:
            Qtn = self.normalizedConeResistance2009(qt,svo,svo_ef,n_init,pa=100)
            Ic = self.SBTindex(Qtn,Rf)
            n = self.stressExponent(Ic,svo_ef)

            if abs(n-n_init) < 0.01:
                break
            else:
                n_init+= 0.01
        return Qtn, Ic, n

    def SBT(self,qt,svo,svo_ef,Rf):
        Qt = self.normalizedConeResistance1990(qt,svo,svo_ef)
        Ic = self.SBTindex(Qt,Rf)
        return Qt, Ic

    def SBT90(self):
        list_Qt = [0]
        list_Ic = [0]
        list_qt1 = [0]

        for i in range(1,len(self.df['qt'])):
            qt = self.df.loc[i,'qt']
            svo = self.df.loc[i,'svo']
            svo_ef = self.df.loc[i,'svo_ef']
            Rf = self.df.loc[i,'Rf']
            Qt, Ic = self.SBT(qt,svo,svo_ef,Rf)
            qt1 = round((qt/100) / (svo_ef/100)**0.5,2)
            list_Qt.append(Qt)
            list_Ic.append(Ic)
            list_qt1.append(qt1)

        self.df['Qt'] = list_Qt
        self.df['Ic_90'] = list_Ic
        self.df['qt1'] = list_qt1

    def SBT2009(self):
        list_Qtn = [0]
        list_Ic = [0]
        list_n = [0]

        for i in range(1,len(self.df['qt'])):
            qt = self.df.loc[i,'qt']
            svo = self.df.loc[i,'svo']
            svo_ef = self.df.loc[i,'svo_ef']
            Rf = self.df.loc[i,'Rf']
            Qtn, Ic, n = self.SBTn(qt,svo,svo_ef,Rf)
            list_Qtn.append(Qtn)
            list_Ic.append(Ic)
            list_n.append(n)

        self.df['Qtn'] = list_Qtn
        self.df['n'] = list_n
        self.df['Ic'] = list_Ic

    def SBTzone(self,Ic):
        if Ic > 3.6:
            zone = 2
        elif Ic > 2.95:
            zone = 3
        elif Ic > 2.6:
            zone = 4
        elif Ic > 2.05:
            zone = 5
        elif Ic > 1.31:
            zone = 6
        else:
            zone = 7
        return int(zone)

    def SBTzone2009(self):
        list_zone = []
        for Ic in self.df['Ic']:
            if Ic > 0:
                zone = self.SBTzone(Ic)
                list_zone.append(zone)
            else:
                zone = None
                list_zone.append(zone)
        self.df['SBTn'] = list_zone
        # self.df['SBTcolor'] = list_zone

    def SBTcolor(self):
        SBTcolor = [None]
        for i, Qtn in enumerate(self.df['Qtn']):
            if i > 0:
                Rf = self.df.loc[i,'Rf']
                if Rf < 0.1:
                    Rf = 0.1
                zone = self.zoneByColor(Rf,Qtn)
                SBTcolor.append(zone)
            else:
                pass
        self.df['SBTcolor'] = SBTcolor

    def calc_Fr(self,Ic,Qtn):
        A = (-2.44 + ((4*Ic**2)+(27.76*log10(Qtn))-(4*log10(Qtn)**2)-48.1636)**0.5)*0.5
        Fr = 10**A
        return Fr

    def calc_Qtn(self,Ic,Fr):
        B = (6.94 + ((4*Ic**2)-(9.76*log10(Qtn))-(4*log10(Qtn)**2)-5.9536)**0.5)*0.5
        Qtn = 10**B
        return Qtn

    def plot_SBTn(self):
        list_Qn3_6 = np.arange(1.45,10,0.01)
        list_Qn2_95 = np.arange(5,53,1)
        list_Qn2_6 = np.arange(8.7,64,1)
        list_Qn2_05 = np.arange(27,175,1)
        list_Qn1_31 = np.arange(150,1000,1)

        list_Qn = [list_Qn3_6,list_Qn2_95,list_Qn2_6,list_Qn2_05,list_Qn1_31]
        list_Ic = [3.6, 2.95, 2.6, 2.05, 1.31]
        list_Fr_zone1 = np.arange(0.1,3,0.1)
        list_Fr_zone_89 = np.arange(1.5,10,0.2)
        fig, axs = plt.subplots(figsize=(5,5.5))

        axs.plot(self.df['Rf'],self.df['Qtn'],marker='o',markeredgecolor='black',markerfacecolor='red',linewidth=0)

        # Zone 1
        list_Qtn_zone1 = []
        for Fr in list_Fr_zone1:
            Qtn_zone1 = 12*exp(-1.4*Fr)
            list_Qtn_zone1.append(Qtn_zone1)
        axs.plot(list_Fr_zone1,list_Qtn_zone1,color='black')

        # Zone 8 & 9
        list_Qn_zone_89 = []
        for Fr in list_Fr_zone_89:
            Qtn = 1/((0.005*(Fr-1))-(0.0003*(Fr-1)**2)-0.002)
            list_Qn_zone_89.append(Qtn)
        axs.plot(list_Fr_zone_89,list_Qn_zone_89,color='black')

        # Zone 2 & 7
        for i, Ic in enumerate(list_Ic):
            list_Rf = []
            Qtn_list = list_Qn[i]
            for Qtn in Qtn_list:
                Rf = self.calc_Fr(Ic,Qtn)
                list_Rf.append(Rf)
            axs.plot(list_Rf,Qtn_list,color='black')

        #Plot zone code
        axs.text(0.3,2,'1',fontsize='large')
        axs.text(6,1.5,'2',fontsize='large')
        axs.text(4,5,'3',fontsize='large')
        axs.text(0.8,8,'4',fontsize='large')
        axs.text(0.25,15,'5',fontsize='large')
        axs.text(1.1,250,'6',fontsize='large')
        axs.text(0.25,400,'7',fontsize='large')
        axs.text(4,400,'8 & 9',fontsize='large')

        note = '1 Sensitive fine-grained\n2 Organic\n3 Clay\n4 Silt-mixtures\n5 Sand-mixtures\n6 Sand\n7 Gravelly sand to sand\n8 Very stiff sand to clayey sand\n9 Very stiff fine-grained'
        axs.annotate(f'Normalized Soil Behaviour Type (SBTn)\n(Robertson, 2009)\n\n{note}',xy = (1.05, 0.1),xycoords='axes fraction')

        axs.set_title(f'{self.id_} - SBTn plot',fontweight='bold')
        axs.set_xlabel(r'Normalized Friction Ratio, $R_r$')
        axs.set_ylabel(r'Normalized Cone Resistance, $Q_{tn}$')
        axs.set_xlim(xmin=0.1,xmax=10)
        axs.set_ylim(ymin=1,ymax=1000)
        axs.set_xscale("log")
        axs.set_yscale("log")

    def frictionAngle(self):
        friction_angle = []
        for i, Qtn in enumerate(self.df['Qtn']):
            if self.df.loc[i,'SBTcolor'] < 5.0 or self.df.loc[i,'SBTcolor'] == 9.0:
                phi = None
            elif i == 0:
                phi = None
            else:
                phi = round(17.6+(11*log10(Qtn)),2)
            friction_angle.append(phi)

        self.df['phi'] = friction_angle

    def undrainedShearStrength(self):
        undrained_shear = []
        for i, qt in enumerate(self.df['qt']):
            if self.df.loc[i,'SBTcolor'] < 5.0 or self.df.loc[i,'SBTcolor'] == 9.0:
                svo = self.df.loc[i,'svo']
                Nkt = 14
                su = round((qt-svo)/Nkt,2)
            elif i == 0:
                su = None
            else:
                su = None
            undrained_shear.append(su)

        self.df['Su'] = undrained_shear

    def OCR_consol(self):
        overconsolidated_ratio = []
        preconsolidation_pressure = []
        exponent_m = []
        insituStress_K0 = []
        meanEffectiveStress = []
        adjustedConeResistance = []
        # exponent_m2 = []
        for i, qt in enumerate(self.df['qt']):
            Ic = self.df.loc[i,'Ic']
            m_ef = round(1- (0.28 / (1 + (Ic/2.65)**25)),2) # Mayne
            exponent_m.append(m_ef)
            zone = self.df.loc[i,'SBTcolor']
            svo = self.df.loc[i,'svo']
            svo_ef = self.df.loc[i,'svo_ef']
            phi = self.df.loc[i,'phi']
            qc = self.df.loc[i,'qt']

            # Mayne 2013 reference
            # if zone == 6 or zone ==7:
            #     m2 = 0.72
            # elif zone == 5:
            #     m2 = 0.8
            # elif zone == 4:
            #     m2 = 0.85
            # else:
            #     m2 = 1
            # exponent_m1.append(m2)

            if i == 0:
                OCR = None
                pc = None
                K0 = None
                pm = None
                qcm = None
                m = None
            else:
                pc = round(self.pc_mayne2012(m_ef,qt,svo),2)
                OCR = round(pc/svo_ef,2)
                K0 = round(self.stressRatio_K0(zone,OCR,phi),2)
                pm = round(svo_ef*((1+(2*K0))/3),2)
                qcm = round(qc*(100/pm)**0.5,2)

            overconsolidated_ratio.append(OCR)
            preconsolidation_pressure.append(pc)
            insituStress_K0.append(K0)
            meanEffectiveStress.append(pm)
            adjustedConeResistance.append(qcm)

        self.df['OCR'] = overconsolidated_ratio
        self.df['pc'] = preconsolidation_pressure
        self.df['K0'] = insituStress_K0
        self.df['m_ef'] = exponent_m
        self.df['pm'] = meanEffectiveStress
        self.df['qcm'] = adjustedConeResistance
        # self.df['m1'] = exponent_m1

    def stressRatio_K0(self,zone,OCR,phi):
        # Robertson 2016 reference
        if zone < 5.0 or zone == 9.0:
            K0 = 0.5*OCR**0.5
        else:
            K0 = (1-np.sin(np.radians(phi)))*OCR**(np.sin(np.radians(phi)))
        return round(K0,3)

    def equivalentN60(self):
        equivalent_N60 = []
        for i, Ic in enumerate(self.df['Ic']):
            if i == 0:
                N60 = None
            else:
                qt = self.df.loc[i,'qt']
                N60 = int((qt/100)/(8.5*(1-(Ic/4.6))))
            equivalent_N60.append(N60)

        self.df['N60'] = equivalent_N60

    def relativeDensity(self):
        relative_density = []
        for i, Qtn in enumerate(self.df['Qtn']):
            if self.df.loc[i,'SBTcolor'] < 4.0 or self.df.loc[i,'SBTcolor'] == 9.0:
                Dr = None
            elif i == 0:
                Dr = None
            else:
                OCR = self.df.loc[i,'OCR']
                qt1 = self.df.loc[i,'qt1']
                # Dr = round(((Qtn/350)**0.5)*100,2) # robertson
                Dr = round(100 * ((qt1/(305*(OCR**0.2)))**0.5), 2) # Mayne
            relative_density.append(Dr)

        self.df['Dr'] = relative_density

    def modulusModifier(self):
        a_value = []
        Massarsch_soiltype = []
        modulusNumber = []
        expoenet_j = []
        for i, zone in enumerate(self.df['SBTcolor']):
            if i == 0:
                a = None
                type = None
                m = None
                j = None
            else:
                compactness = self.df.loc[i,'compactness']
                if zone == 7:
                    # Gravel
                    j = 1
                    if compactness == 'dense' or compactness == 'very dense':
                        a = 45
                        type = 'Gravel, dense'
                    elif compactness == 'medium':
                        a = 40
                        type = 'Gravel, compact'
                    else:
                        a = 35
                        type = 'Gravel, loose'
                elif zone == 6 or zone == 5:
                    # Sand
                    if compactness == 'dense' or compactness == 'very dense':
                        a = 35
                        type = 'Sand, dense'
                        j = 1
                    elif compactness == 'medium':
                        a = 28 #Ok 40; Agus
                        type = 'Sand, compact'
                        j = 0.75
                    else:
                        j = 0.5
                        if zone == 5:
                            a = 20
                            type = 'Sand, silty loose'
                        else:
                            a = 22
                            type = 'Sand, loose'
                elif zone == 4:
                    # Silt
                    if compactness == 'dense' or compactness == 'very dense':
                        a = 20
                        type = 'Silt, dense'
                        j = 1
                    elif compactness == 'medium':
                        a = 15
                        type = 'Silt, compact'
                        j = 1
                    else:
                        a = 12
                        type = 'Silt, loose'
                        j = 0.5
                elif zone == 2:
                    a = 7 # organic
                    type = 'Organic, soft'
                    j = 0.01
                else:
                    # Clay
                    j = 0.01
                    if compactness == 'stiff' or compactness == 'very stiff':
                        a = 5
                        type = 'Clay, firm'
                    else:
                        a = 3
                        type = 'Clay, soft'

                qcm = self.df.loc[i,'qcm']
                pr = 100
                m = round(a*(qcm/pr)**0.5,2)
            a_value.append(a)
            modulusNumber.append(m)
            Massarsch_soiltype.append(type)
            expoenet_j.append(j)

        self.df['a'] = a_value
        self.df['soil type'] = Massarsch_soiltype
        self.df['m'] = modulusNumber
        self.df['j'] = expoenet_j

    def density(self,Dr):
        # Budhu; reference
        if Dr < 20:
            density = 'very loose'
        elif Dr < 40:
            density = 'loose'
        elif Dr < 70:
            density = 'medium'
        elif Dr < 85:
            density = 'dense'
        else:
            density = 'very dense'
        return density

    def consistency(self,Su):
        # Budhu; reference
        if Su < 10:
            cons = 'very soft'
        elif Su < 25:
            cons = 'soft'
        elif Su < 50:
            cons = 'medium'
        elif Su < 100:
            cons = 'stiff'
        elif Su < 200:
            cons = 'very stiff'
        else:
            cons = 'extremly stiff'
        return cons

    def compactness(self):
        compactness = []
        for i, Dr in enumerate(self.df['Dr']):
            Su = self.df.loc[i,'Su']
            if i == 0:
                compactness.append(None)
            else:
                if type(Dr) == float:
                    compactness.append(self.density(Dr))
                else:
                    compactness.append(self.consistency(Su))
        self.df['compactness'] = compactness

    def constrainModulus(self):
        constraint_modulus = []
        for i, qt in enumerate(self.df['qt']):
            if i<0:
                M = None
            else:
                Ic = self.df.loc[i,'Ic']
                Qt = self.df.loc[i,'Qt']
                qt = self.df.loc[i,'qt']
                svo = self.df.loc[i,'svo']
                if Ic > 2.2:
                    if Qt<14:
                        am = Qt
                    else:
                        am = 14
                else:
                    am = 0.0188*(10**((0.55*Ic)+1.68))
                M = round(am*(qt-svo),0)
            constraint_modulus.append(M)

        self.df['M'] = constraint_modulus

    def youngsModulus(self):
        youngs_modulus = []
        for i, qt in enumerate(self.df['qt']):
            Ic = self.df.loc[i,'Ic']
            if i<0 or Ic>2.6:
                E = None
            else:
                qt = self.df.loc[i,'qt']
                svo = self.df.loc[i,'svo']
                am = 0.015*(10**((0.55*Ic)+1.68))
                E = round(am*(qt-svo),0)
            youngs_modulus.append(E)

        self.df['E'] = youngs_modulus

    def smallShearModulus(self):
        small_shear_modulus = []
        for i, qt in enumerate(self.df['qt']):
            Ic = self.df.loc[i,'Ic']
            if i<0:
                G0 = None
            else:
                qt = self.df.loc[i,'qt']
                svo = self.df.loc[i,'svo']
                am = 0.0188*(10**((0.55*Ic)+1.68))
                G0 = round(am*(qt-svo),0)
            small_shear_modulus.append(G0)

        self.df['G0'] = small_shear_modulus

    def shearWaveVelocity(self):
        shear_wave_velocity = []
        for i, G0 in enumerate(self.df['G0']):
            if i<0:
                Vs = None
            else:
                g = self.df.loc[i,'g']
                Vs = round(sqrt(G0/(g/9.81)),2)
            shear_wave_velocity.append(Vs)

        self.df['Vs'] = shear_wave_velocity

    def rigidityIndex(self):
        small_strain_rigidity = []
        normalized_small_strain_rigidity = []
        for i, G0 in enumerate(self.df['G0']):
            if i<1:
                Ig = None
                Kg = None
            else:
                qt = self.df.loc[i,'qt']
                svo = self.df.loc[i,'svo']
                Qtn = self.df.loc[i,'Qtn']
                qn = qt-svo
                Ig = round(G0/qn,2)
                Kg = round(Ig*(Qtn**0.75),2)
            small_strain_rigidity.append(Ig)
            normalized_small_strain_rigidity.append(Kg)

        self.df['IG'] = small_strain_rigidity
        self.df['KG'] = normalized_small_strain_rigidity

    def solve_settlement(self,foundation='circular',q=0,B=0,Df=0):
        list_pz = []
        # list_strain = []
        list_settlement = []
        for i, z in enumerate(self.df['z']):
            if i == 0:
                strain_e = 0
                pz = 0
                Sz = 0
            else:
                if z > B or z < Df:
                    strain_e = 0
                    pz = 0
                    Sz = 0
                else:
                    if foundation == 'circular':
                        pz = pondasi.Stresses.circular(z,B,q)
                    else:
                        pz = pondasi.Stresses.rect(z,B,B,q)
                    pr = 100 # kPa
                    svo_ef = self.df.loc[i,'svo_ef']
                    m = self.df.loc[i,'m']
                    j = self.df.loc[i,'j']
                    dz = 0.2 # m; only for sondir
                    p1 = pz + svo_ef
                    M = m*j
                    P0 = (svo_ef/pr)**j
                    P1 = (p1/pr)**j
                    strain_e = (1/M)*(P1-P0)
                    Sz = strain_e*dz

            # list_strain.append(strain_e)
            list_pz.append(pz)
            list_settlement.append(Sz*1000)

        # self.df['strain'] = list_strain
        self.df['pz'] = list_pz
        self.df['Sz'] = list_settlement

        cumulative_Sz = []
        for i, z in enumerate(self.df['z']):
            if z < Df or z > B:
                Sz_cum = 0
            else:
                Sz_cum = sum(self.df.loc[i:,'Sz'])
            cumulative_Sz.append(Sz_cum)
        self.df['Sz_tot'] = cumulative_Sz

    def qa_CPT(self,B=1,L=0,Df=1):
        if L == 0:
            L = B

        i_start = int((Df+(B/2))/0.2)
        # i_start = int(Df/0.2)
        i_end = int((B+Df)/0.2)

        dz = 0.2
        qt_dz = sum((self.df.loc[i_start:i_end,'qt']*dz)) # kPa
        z_tot = B
        self.qt_avg = qt_dz/z_tot

        if B < 1.2:
            qa = (self.qt_avg/30)
        else:
            qa = (self.qt_avg/50) * ((B+0.3)/B)**2
        self.qa_cpt = round(qa,2)

    def plot_rigidityIndex(self):
        list_Qtn_Ig = np.arange(1,1000,50)
        fig, axs = plt.subplots(figsize=(5,5.5))
        axs.plot(self.df['IG'],self.df['Qtn'],marker='o',markeredgecolor='black',markerfacecolor='red',linewidth=0)

        list_Qtn_KG100 = []
        list_Qtn_KG330 = []
        for Qtn in list_Qtn_Ig:
            Ig_100 = 100/(Qtn**0.75)
            Ig_330 = 330/(Qtn**0.75)
            list_Qtn_KG100.append(Ig_100)
            list_Qtn_KG330.append(Ig_330)

        axs.plot(list_Qtn_KG100,list_Qtn_Ig,color='black')
        axs.plot(list_Qtn_KG330,list_Qtn_Ig,color='black')

        #Plot zone code
        axs.text(80,1.5,f'young\nuncemeneted',fontsize='medium',ha='center',rotation=-62)
        axs.text(20,1.5,r'$K*_G = 100$',fontsize='medium',rotation=-62)
        axs.text(130,1.5,r'$K*_G = 330$',fontsize='medium',rotation=-62)
        axs.text(100,300,f'Soil with microstructure\n(e.g. cementation/bonding\nand aging)',fontsize='medium',ha='center')

        axs.set_title(f'{self.id_} - Normalized Rigidity Index',fontweight='bold')
        axs.set_xlabel(r'Small-strain Rigidity Index, $I_G = G_0/q_n$')
        axs.set_ylabel(r'Normalized Cone Resistance, $Q_{tn}$')
        axs.set_xlim(xmin=0.1,xmax=1000)
        axs.set_ylim(ymin=1,ymax=1000)
        # axs.grid(True,which='major')
        # axs.grid(True,which='minor')
        axs.set_xscale("log")
        axs.set_yscale("log")

    def solve_basic(self):
        self.average_qc()
        self.average_fs()
        self.solve_unitWeight()
        self.solve_RfSvo()
        self.solve_porePressure()
        self.SBT90()
        self.SBT2009()
        self.SBTzone2009()
        self.SBTcolor()
        self.smallShearModulus()
        self.rigidityIndex()

    def solve_parameters(self):
        self.frictionAngle()
        self.undrainedShearStrength()
        self.OCR_consol()
        self.equivalentN60()
        self.relativeDensity()
        self.compactness()
        self.modulusModifier()
        self.constrainModulus()
        self.youngsModulus()
        self.shearWaveVelocity()

    def raw_plot(self):
        f, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(6,7),sharey=True)
        ax1.plot(self.df['qc'],self.df['z'],color='black')
        ax1.set_xlabel(r'$q_c$ ($Kg/cm^2$)')
        ax1.set_ylabel(r'Depth, $z$ (m)')
        # ax1.set_xticks(np.arange(0, 50, 10))
        # ax1.set_xticks(np.arange(0, 50, 5),minor=True)
        # ax1.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax1.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        # ax1.set_ylim(0,self.end_depth)
        ax1.invert_yaxis()
        ax1.grid(True,which='major')
        ax1.grid(True,which='minor',linewidth=0.1)
        ax1.set_title('Cone resistance',fontweight='bold',fontsize='small')

        ax2.plot(self.df['fs'],self.df['z'],color='black')
        ax2.set_xlabel(r'$f_s$ ($Kg/cm^2$)')
        ax2.set_ylabel(r'Depth, $z$ (m)')
        # ax2.set_xticks(np.arange(0, 1000, 200))
        # ax2.set_xticks(np.arange(0, 1000, 50),minor=True)
        # ax2.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax2.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        ax2.grid(True,which='major')
        ax2.grid(True,which='minor',linewidth=0.1)
        ax2.set_title('Sleeve friction',fontweight='bold',fontsize='small')

        ax3.plot(self.df['u'],self.df['z'],color='blue')
        ax3.set_xlabel(r'$u$ (kPa)')
        ax3.set_ylabel(r'Depth, $z$ (m)')
        # ax3.set_xticks(np.arange(0, 200, 50))
        # ax3.set_xticks(np.arange(0, 200, 10),minor=True)
        # ax3.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax3.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        ax3.grid(True,which='major')
        ax3.grid(True,which='minor',linewidth=0.1)
        ax3.set_title('Pore pressure',fontweight='bold',fontsize='small')
        plt.show()

    def basic_plot(self):
        f, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(6,7),sharey=True)
        ax1.plot(self.df['qt']/1000,self.df['z'],color='black')
        ax1.set_xlabel(r'$q_t$ (MPa)')
        ax1.set_ylabel(r'Depth, $z$ (m)')
        # ax1.set_xticks(np.arange(0, 30, 5))
        # ax1.set_xticks(np.arange(0, 30, 1),minor=True)
        # ax1.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax1.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        # ax1.set_ylim(0,self.end_depth)
        ax1.invert_yaxis()
        ax1.grid(True,which='major')
        ax1.grid(True,which='minor',linewidth=0.1)
        ax1.set_title('Cone resistance qt',fontweight='bold',fontsize='small')

        ax2.plot(self.df['Rf'],self.df['z'],color='black')
        ax2.set_xlabel(r'$R_f$ (%)')
        ax2.set_ylabel(r'Depth, $z$ (m)')
        # ax2.set_xticks(np.arange(0, 10, 2))
        # ax2.set_xticks(np.arange(0, 10, 1),minor=True)
        # ax2.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax2.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        ax2.grid(True,which='major')
        ax2.grid(True,which='minor',linewidth=0.1)
        ax2.set_title('Friction ratio',fontweight='bold',fontsize='small')

        ax3.plot(self.df['u'],self.df['z'],color='blue')
        ax3.set_xlabel(r'$u$ (kPa)')
        ax3.set_ylabel(r'Depth, $z$ (m)')
        # ax3.set_xticks(np.arange(0, 200, 50))
        # ax3.set_xticks(np.arange(0, 200, 10),minor=True)
        # ax3.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax3.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        ax3.grid(True,which='major')
        ax3.grid(True,which='minor',linewidth=0.1)
        ax3.set_title('Pore pressure',fontweight='bold',fontsize='small')
        plt.show()

    def export_csv(self,name='data export',folder='data_csv',existing_folder=False):
        if existing_folder == True:
            pass
        else:
            os.mkdir(folder)
        self.df.to_csv(f'{folder}\{name}.csv')

    def color_func(self,val):
        if val == 1:
            color = (0.8706,0.8706,0.8706)
        elif val == 2:
            color = (0.7098,0.4118,0.2471)
        elif val == 3:
            color = (0.4,0.4824,0.702)
        elif val == 4:
            color = (0.2706,0.3255,0.5098)
        elif val == 5:
            color = (0.2667,0.5804,0.5176)
        elif val == 6:
            color = (1,0.9725,0)
        elif val == 7:
            color = (0.9765,0.6275,0.3059)
        elif val == 8:
            color = (0.5843,0.5843,0.5843)
        else:
            color = (0.8667,0.1294,0.098)
        return color

    def soil_profil(self):
        # Create new color column with map()
        self.df['colors'] = self.df['SBTcolor'].apply(self.color_func)
        f, ax = plt.subplots(figsize=(2,7),sharey=True)
        #--------------------------------------------------------------------------
        self.df.plot('z','SBTcolor', kind='barh',ax=ax,color=self.df['colors'],width=1)

        ax.set_xlabel(r'SBTn')
        ax.set_ylabel(r'Depth, $z$ (m)')

        N_data = len(self.df['z']+1)
        # ax.set_xticks(np.arange(0, 10,1))
        ax.set_yticks(np.arange(0, N_data, 5))
        ax.yaxis.set_major_formatter(lambda x, pos: str(x/5))
        ax.set_title('Soil Profil',fontweight='bold',fontsize='medium')
        ax = plt.gca()
        ax.invert_yaxis()
        ax.get_legend().remove()

        z1 = mpatches.Patch(color=(0.8706,0.8706,0.8706), label='1. Sensitive fine-grained')
        z2 = mpatches.Patch(color=(0.7098,0.4118,0.2471), label='2. Organic')
        z3 = mpatches.Patch(color=(0.4,0.4824,0.702), label='3. Clay')
        z4 = mpatches.Patch(color=(0.2706,0.3255,0.5098), label='4. Silt-mixtures')
        z5 = mpatches.Patch(color=(0.2667,0.5804,0.5176), label='5. Sand-mixtures')
        z6 = mpatches.Patch(color=(1,0.9725,0), label='6. Sand')
        z7 = mpatches.Patch(color=(0.9765,0.6275,0.3059), label='7. Gravelly sand to sand')
        z8 = mpatches.Patch(color=(0.5843,0.5843,0.5843), label='8. Very stiff sand to clayey sand')
        z9 = mpatches.Patch(color=(0.8667,0.1294,0.098), label='9. Very stiff fine-grained')
        plt.legend(handles=[z1,z2,z3,z4,z5,z6,z7,z8,z9],bbox_to_anchor =(1.1, 0.9),fontsize='small')

    def plot_norm(self):
        self.df.loc[0,'Ic'] = None
        f, (ax1, ax2) = plt.subplots(1,2,figsize=(4,7),sharey=True)
        #-------------------------------------------------------------------------
        ax1.plot(self.df['Rf'],self.df['z'],color='black')
        ax1.set_xlabel(r'$R_f$')
        ax1.set_ylabel(r'Depth, $z$ (m)')
        # ax1.set_xticks(np.arange(0, 15, 5))
        # ax1.set_xticks(np.arange(0, 15, 1),minor=True)
        # ax1.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax1.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        ax1.grid(True,which='major')
        ax1.grid(True,which='minor',linewidth=0.1)
        # ax1.set_ylim(0,self.end_depth)
        ax1.invert_yaxis()
        ax1.set_title('Friction Ratio Rf',fontweight='bold',fontsize='small')

        #---------------------------------------------------------------------------
        ax2.plot(self.df['Ic'],self.df['z'],color='black')
        z = self.end_depth-2
        ax2.axvspan(z,4,facecolor=(0.7098,0.4118,0.2471))
        ax2.axvspan(z,3.6,facecolor=(0.4,0.4824,0.702))
        ax2.axvspan(z,2.95,facecolor=(0.2706,0.3255,0.5098))
        ax2.axvspan(z,2.6,facecolor=(0.2667,0.5804,0.5176))
        ax2.axvspan(z,2.05,facecolor=(1,0.9725,0))
        ax2.axvspan(z,1.31,facecolor=(0.9765,0.6275,0.3059))
        ax2.set_xlabel(r'$I_c$')
        ax2.set_ylabel(r'Depth, $z$ (m)')
        ax2.set_xlim(xmin=1,xmax=4)
        ax2.set_title('SBTn Index',fontweight='bold',fontsize='small')

        z1 = mpatches.Patch(color=(0.8706,0.8706,0.8706), label='1. Sensitive fine-grained (N/A)')
        z2 = mpatches.Patch(color=(0.7098,0.4118,0.2471), label='2. Organic (Ic>4.6)')
        z3 = mpatches.Patch(color=(0.4,0.4824,0.702), label='3. Clay (2.95-3.6)')
        z4 = mpatches.Patch(color=(0.2706,0.3255,0.5098), label='4. Silt-mixtures (2.60-2.95)')
        z5 = mpatches.Patch(color=(0.2667,0.5804,0.5176), label='5. Sand-mixtures (2.05-2.60)')
        z6 = mpatches.Patch(color=(1,0.9725,0), label='6. Sand (1.31-2.05)')
        z7 = mpatches.Patch(color=(0.9765,0.6275,0.3059), label='7. Gravelly sand to sand (Ic<1.31)')
        z8 = mpatches.Patch(color=(0.5843,0.5843,0.5843), label='8. Very stiff sand to clayey sand (N/A)')
        z9 = mpatches.Patch(color=(0.8667,0.1294,0.098), label='9. Very stiff fine-grained (N/A)')
        plt.legend(handles=[z1,z2,z3,z4,z5,z6,z7,z8,z9],bbox_to_anchor =(1.1, 0.9),fontsize='small')

    def estimation_plot1(self):
        f, (ax1, ax2, ax3, ax4) = plt.subplots(1, 4, figsize=(8,7),sharey=True)
        ax1.plot(self.df['g'],self.df['z'],color='black')
        ax1.set_xlabel(r'$\gamma_s\ (kN/m^3)$')
        ax1.set_ylabel(r'Depth, $z$ (m)')
        ax1.set_xticks(np.arange(15, 20, 1))
        ax1.set_xticks(np.arange(15, 20, 0.5),minor=True)
        # ax1.set_yticks(np.arange(0, self.end_depth, -2))
        # ax1.set_yticks(np.arange(0, self.end_depth, -0.5),minor=True)
        ax1.grid(True,which='major')
        ax1.grid(True,which='minor',linewidth=0.1)
        # ax1.set_ylim(0,self.end_depth)
        ax1.invert_yaxis()
        ax1.set_title('Unit weight',fontweight='bold',fontsize='small')

        ax2.plot(self.df['phi'],self.df['z'],color='black')
        ax2.set_xlabel(r'$\phi$ (deg)')
        ax2.set_ylabel(r'Depth, $z$ (m)')
        ax2.set_xticks(np.arange(25, 45, 5))
        ax2.set_xticks(np.arange(25, 45, 2.5),minor=True)
        # ax2.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax2.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        ax2.grid(True,which='major')
        ax2.grid(True,which='minor',linewidth=0.1)
        ax2.set_title('Int. friction angle',fontweight='bold',fontsize='small')

        ax3.plot(self.df['Su'],self.df['z'],color='black')
        ax3.set_xlabel(r'$S_u$ (kPa)')
        ax3.set_ylabel(r'Depth, $z$ (m)')
        # ax3.set_xticks(np.arange(0, 200, 50))
        # ax3.set_xticks(np.arange(0, 200, 10),minor=True)
        # ax3.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax3.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        ax3.grid(True,which='major')
        ax3.grid(True,which='minor',linewidth=0.1)
        ax3.set_title('Und. shear strength',fontweight='bold',fontsize='small')

        ax4.plot(self.df['N60'],self.df['z'],color='black')
        ax4.set_xlabel(r'$N_{60}$')
        ax4.set_ylabel(r'Depth, $z$ (m)')
        ax4.set_xticks(np.arange(0, 50, 10))
        ax4.set_xticks(np.arange(0, 50, 5),minor=True)
        # ax4.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax4.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        ax4.grid(True,which='major')
        ax4.grid(True,which='minor',linewidth=0.1)
        ax4.set_title('Equivalent $N_{60}$',fontweight='bold',fontsize='small')
        plt.show()

    def estimation_plot2(self):
        f, (ax1, ax2, ax3, ax4, ax5) = plt.subplots(1, 5, figsize=(10,7),sharey=True)
        ax1.plot(self.df['Dr'],self.df['z'],color='black')
        ax1.set_xlabel(r'$D_r$ (%)')
        ax1.set_ylabel(r'Depth, $z$ (m)')
        ax1.set_xticks(np.arange(0, 80, 20))
        ax1.set_xticks(np.arange(0, 80, 5),minor=True)
        # ax1.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax1.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        # ax1.set_ylim(0,self.end_depth)
        ax1.invert_yaxis()
        ax1.grid(True,which='major')
        ax1.grid(True,which='minor',linewidth=0.1)
        ax1.set_title('Relative density',fontweight='bold',fontsize='small')

        ax2.plot(self.df['OCR'],self.df['z'],color='black')
        ax2.set_xlabel(r'OCR')
        ax2.set_ylabel(r'Depth, $z$ (m)')
        # ax2.set_xticks(np.arange(0, 20, 5))
        # ax2.set_xticks(np.arange(0, 20, 1),minor=True)
        # ax2.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax2.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        ax2.grid(True,which='major')
        ax2.grid(True,which='minor',linewidth=0.1)
        ax2.set_title('Over consol. ratio',fontweight='bold',fontsize='small')

        ax3.plot(self.df['M'],self.df['z'],color='black')
        ax3.set_xlabel(r'M (kPa)')
        ax3.set_ylabel(r'Depth, $z$ (m)')
        ax3.set_xticks(np.arange(0, 200000, 50000))
        ax3.ticklabel_format(axis="x", style="sci", scilimits=(0,0))
        ax3.set_xticks(np.arange(0, 200000, 10000),minor=True)
        # ax3.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax3.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        ax3.grid(True,which='major')
        ax3.grid(True,which='minor',linewidth=0.1)
        ax3.set_title('Constrain modulus',fontweight='bold',fontsize='small')

        ax4.plot(self.df['E'],self.df['z'],color='black')
        ax4.set_xlabel(r'E (kPa)')
        ax4.set_ylabel(r'Depth, $z$ (m)')
        ax4.set_xticks(np.arange(0, 200000, 50000))
        ax4.ticklabel_format(axis="x", style="sci", scilimits=(0,0))
        ax4.set_xticks(np.arange(0, 200000, 10000),minor=True)
        # ax4.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax4.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        ax4.grid(True,which='major')
        ax4.grid(True,which='minor',linewidth=0.1)
        ax4.set_title('Youngs modulus',fontweight='bold',fontsize='small')

        ax5.plot(self.df['G0'],self.df['z'],color='black')
        ax5.set_xlabel(r'$G_0$ (kPa)')
        ax5.set_ylabel(r'Depth, $z$ (m)')
        ax5.set_xticks(np.arange(0, 200000, 50000))
        ax5.ticklabel_format(axis="x", style="sci", scilimits=(0,0))
        ax5.set_xticks(np.arange(0, 200000, 10000),minor=True)
        # ax5.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax5.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        ax5.grid(True,which='major')
        ax5.grid(True,which='minor',linewidth=0.1)
        ax5.set_title('Small-strain modulus',fontweight='bold',fontsize='small')
        plt.show()

    def estimation_plot3(self):
        f, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(6,7),sharey=True)
        ax1.plot(self.df['Vs'],self.df['z'],color='black')
        ax1.set_xlabel(r'$V_s$ (m/s)')
        ax1.set_ylabel(r'Depth, $z$ (m)')
        ax1.set_xticks(np.arange(0, 300, 100))
        ax1.set_xticks(np.arange(0, 300, 50),minor=True)
        # ax1.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax1.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        ax1.grid(True,which='major')
        ax1.grid(True,which='minor',linewidth=0.1)
        # ax1.set_ylim(0,self.end_depth)
        ax1.invert_yaxis()
        ax1.set_title('Shear wave velocity',fontweight='bold',fontsize='small')

        ax2.plot(self.df['IG'],self.df['z'],color='black')
        ax2.set_xlabel(r'$I_G$')
        ax2.set_ylabel(r'Depth, $z$ (m)')
        ax2.set_xticks(np.arange(0, 60, 20))
        ax2.set_xticks(np.arange(0, 60, 5),minor=True)
        # ax2.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax2.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        ax2.grid(True,which='major')
        ax2.grid(True,which='minor',linewidth=0.1)
        ax2.set_title('Rigidity index',fontweight='bold',fontsize='small')

        ax3.plot(self.df['KG'],self.df['z'],color='black')
        ax3.set_xlabel(r'$K_G$')
        ax3.set_ylabel(r'Depth, $z$ (m)')
        ax3.set_xticks(np.arange(100, 500, 100))
        ax3.set_xticks(np.arange(100, 500, 50),minor=True)
        # ax3.set_yticks(np.arange(0, self.end_depth-2, -2))
        # ax3.set_yticks(np.arange(0, self.end_depth-2, -0.5),minor=True)
        ax3.grid(True,which='major')
        ax3.grid(True,which='minor',linewidth=0.1)
        ax3.set_title('Normalized rigidity ind.',fontweight='bold',fontsize='small')
        plt.show()

class Segment:
    """docstring for ."""

    def __init__(self, id_='Segment 1'):
        self.id_ = id_

    def read_excel(self, name, colz='A',colqc='F',colfs='G',data_row=15,cell_date='D5', cell_loc='D6', cell_GWL='J7', cell_coord='D7'):
        workbook = load_workbook(filename=name,data_only=True)
        sheets = workbook.sheetnames
        self.raw_df = []
        self.sheet_name = []
        self.list_date = []
        self.list_loc = []
        self.list_GWL = []
        self.list_coord = []

        for name in sheets:
            list_z = [0]
            list_qc = [0]
            list_fs = [0]

            if name[0:4] == 'DATA':
                self.sheet_name.append(name)
                sheet = workbook[name]
                self.list_date.append(sheet[cell_date].value)
                self.list_loc.append(sheet[cell_loc].value)
                self.list_GWL.append(sheet[cell_GWL].value)
                self.list_coord.append(sheet[cell_coord].value)
                i = data_row
                while True:
                    qc = sheet[f'{colqc}{i}'].value
                    if not type(qc) is float:
                        data = list(zip(list_z,list_qc,list_fs))
                        df = pd.DataFrame(data,columns = ['z', 'qc','fs'])
                        df['u2'] = 0
                        self.raw_df.append(df)
                        break
                    elif qc > 0:
                        list_z.append(sheet[f'{colz}{i}'].value)
                        list_qc.append(sheet[f'{colqc}{i}'].value)
                        list_fs.append(sheet[f'{colfs}{i}'].value)
                        i+=1
            else:
                pass

    def plot_qc(self):
        Ngraph = len(self.raw_df)
        b = 2*Ngraph
        f, ax = plt.subplots(1,Ngraph,figsize=(b,6),sharey=True)
        ax[0].invert_yaxis()
        plt.subplots_adjust(left=0.1,bottom=0.1,right=0.9,top=0.9,wspace=0.4,hspace=0.4)
        for i, df in enumerate(self.raw_df):
            ax[i].plot(df['qc'],df['z'])
            ax[i].set_xlabel(f'$q_c$ $Kg/cm^2$')
            ax[i].set_ylabel(f'Depth, z (m)')
            ax[i].set_title(f'CPT {i+1}',fontweight='bold',fontsize='medium')

    def plot_fs(self):
        Ngraph = len(self.raw_df)
        b = 2*Ngraph
        f, ax = plt.subplots(1,Ngraph,figsize=(b,6),sharey=True)
        plt.subplots_adjust(left=0.1,bottom=0.1,right=0.9,top=0.9,wspace=0.4,hspace=0.4)
        ax[0].invert_yaxis()
        for i, df in enumerate(self.raw_df):
            ax[i].plot(df['fs'],df['z'])
            ax[i].set_xlabel(f'$f_s$ $Kg/cm^2$')
            ax[i].set_ylabel(f'Depth, z (m)')
            ax[i].set_title(f'CPT {i+1}',fontweight='bold',fontsize='medium')

    def color_func(self,val):
        if val == 1:
            color = (0.8706,0.8706,0.8706)
        elif val == 2:
            color = (0.7098,0.4118,0.2471)
        elif val == 3:
            color = (0.4,0.4824,0.702)
        elif val == 4:
            color = (0.2706,0.3255,0.5098)
        elif val == 5:
            color = (0.2667,0.5804,0.5176)
        elif val == 6:
            color = (1,0.9725,0)
        elif val == 7:
            color = (0.9765,0.6275,0.3059)
        elif val == 8:
            color = (0.5843,0.5843,0.5843)
        else:
            color = (0.8667,0.1294,0.098)
        return color

    def run(self):
        self.list_df = []
        for i, df in enumerate(self.raw_df):
            s1 = Single(id_=f'CPT{i+1}',qc_unit='Kg/cm2',fs_unit='Kg/cm2')
            s1.data_frame(df)
            GWL = self.list_GWL[i]
            s1.set_date(self.list_date[i])
            s1.set_location(self.list_loc[i])
            s1.set_coordinate(self.list_coord[i])
            if GWL == '-' or GWL == None or isinstance(GWL, str):
                s1.setGWL()
            else:
                s1.setGWL(z=GWL)
            s1.solve_basic()
            s1.solve_parameters()
            self.list_df.append(s1)

    def plot_profil(self):
        daftar_z = []
        for i, s in enumerate(self.list_df):
            daftar_z.append(max(s.df['z']))

        N_data=int(np.ceil(max(daftar_z))/0.2)
        Ngraph = len(self.list_df)
        b = 2*Ngraph

        f, ax = plt.subplots(1,Ngraph,figsize=(b,6))
        for i, s in enumerate(self.list_df):
            df = s.df
            df['colors'] = df['SBTcolor'].apply(self.color_func)

            #--------------------------------------------------------------------------
            df.plot('z','SBTcolor', kind='barh',ax=ax[i],color=df['colors'],width=1)
            ax[i].set_ylim(ymin=0,ymax=N_data)
            ax[i].axes.xaxis.set_visible(False)
            ax[i].get_legend().remove()
            ax[i].invert_yaxis()

            if i == 0:
                ax[i].set_ylabel(r'Depth, $z$ (m)')
            else:
                ax[i].yaxis.label.set_color('w')
            ax[i].set_yticks(np.arange(0, N_data, 5))
            ax[i].yaxis.set_major_formatter(lambda x, pos: str(x/5))
            ax[i].set_title(f'{self.sheet_name[i]}',fontweight='bold',fontsize='medium')

            plt.subplots_adjust(left=0.1,
                            bottom=0.1,
                            right=0.9,
                            top=0.9,
                            wspace=0.4,
                            hspace=0.4)

            if i == len(self.list_df)-1:
                z1 = mpatches.Patch(color=(0.8706,0.8706,0.8706), label='1. Sensitive fine-grained')
                z2 = mpatches.Patch(color=(0.7098,0.4118,0.2471), label='2. Organic')
                z3 = mpatches.Patch(color=(0.4,0.4824,0.702), label='3. Clay')
                z4 = mpatches.Patch(color=(0.2706,0.3255,0.5098), label='4. Silt-mixtures')
                z5 = mpatches.Patch(color=(0.2667,0.5804,0.5176), label='5. Sand-mixtures')
                z6 = mpatches.Patch(color=(1,0.9725,0), label='6. Sand')
                z7 = mpatches.Patch(color=(0.9765,0.6275,0.3059), label='7. Gravelly sand to sand')
                z8 = mpatches.Patch(color=(0.5843,0.5843,0.5843), label='8. Very stiff sand to clayey sand')
                z9 = mpatches.Patch(color=(0.8667,0.1294,0.098), label='9. Very stiff fine-grained')
                plt.legend(handles=[z1,z2,z3,z4,z5,z6,z7,z8,z9],bbox_to_anchor =(1.1, 0.9),fontsize='small')

    def export_csv(self,folder='data_csv',existing_folder=True):
        for i, s in enumerate(self.list_df):
            s.export_csv(name=self.sheet_name[i],folder=folder, existing_folder=existing_folder)

class MSegment:
    """docstring for ."""

    def __init__(self, id_='Multi segment'):
        self.id_ = id_

    def sheet_name(self,name='data',N=15,colz='A',colqc='F',colfs='G',data_row=15,cell_date='D5', cell_loc='D6', cell_GWL='J7', cell_coord='D7'):
        self.previx_name = name
        self.Nfile = N
        self.list_sg = []
        for i in range(1,self.Nfile+1):
            print('Analyze segment',i)
            segmen = Segment()
            segmen.read_excel(f"{self.previx_name}{i}.xlsx",colz,colqc,colfs,data_row,cell_date, cell_loc, cell_GWL, cell_coord)
            # segmen.plot_qc()
            # segmen.plot_fs()
            segmen.run()
            self.list_sg.append(segmen)
            if i == self.Nfile:
                print("Finish!")

    def plot_profil(self):
        for sg in self.list_sg:
            sg.plot_profil()

    def export_csv(self,name='DATA',folder='data_csv',existing_folder=True):
        for sg in self.list_sg:
            for i, s in enumerate(sg.list_df):
                s.export_csv(name=f'{sg.sheet_name[i]}',folder=folder, existing_folder=existing_folder)
